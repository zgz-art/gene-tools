import openpyxl
from openpyxl.styles import numbers
from copy import copy
from datetime import datetime
from .helper import normalize_date

def find_cell_by_value(worksheet, value, exact=False):
    """遍历查找单元格（支持模糊匹配或精确匹配），返回 (row, col) 或 None"""
    for row in worksheet.iter_rows():
        for cell in row:
            cell_val = cell.value
            if cell_val is None:
                continue
            if exact:
                if cell_val == value:
                    return cell.row, cell.column
            else:
                if value.lower() in str(cell_val).lower():
                    return cell.row, cell.column
    return None

def find_row_with_keyword(worksheet, keyword):
    """返回包含关键字的行索引（行号从1开始）"""
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            val = worksheet.cell(row, col).value
            if val and keyword in str(val):
                return row
    return None

def copy_row_style(source_row, target_row, worksheet, max_col):
    """复制源行样式到目标行（用于新增行时保留格式）"""
    for col in range(1, max_col + 1):
        source_cell = worksheet.cell(source_row, col)
        target_cell = worksheet.cell(target_row, col)
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.alignment = copy(source_cell.alignment)

def insert_rows(worksheet, start_row, count, max_col):
    """在 start_row 之前插入 count 行，并复制 start_row 的样式"""
    worksheet.insert_rows(start_row, count)
    # 为新插入的行复制样式（从原来的第 start_row 行复制）
    for i in range(count):
        new_row = start_row + i
        copy_row_style(start_row + count, new_row, worksheet, max_col)

def clear_rows(worksheet, start_row, end_row, max_col):
    """清空指定行区域的内容（保留样式）"""
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            worksheet.cell(row, col).value = None

def fill_basic_info(worksheet, basic_data):
    """模糊匹配基础信息字段，填充到右侧单元格（假设字段在A列，右侧B列）"""
    for key, value in basic_data.items():
        if not value:
            continue
        # 模糊查找包含 key 的单元格
        pos = find_cell_by_value(worksheet, key, exact=False)
        if pos:
            row, col = pos
            # 右侧一格
            target_cell = worksheet.cell(row, col + 1)
            target_cell.value = value

def fill_education_block(worksheet, keyword, edu_data):
    """
    精确匹配 keyword（如“本科学历”）所在行，然后处理接下来的若干行。
    假设格式：A列为字段名，B列为待填充值。
    edu_data 为字典，字段名与模板中的文本进行精确匹配（忽略空格）。
    """
    pos = find_cell_by_value(worksheet, keyword, exact=True)
    if not pos:
        return
    row_start = pos[0] + 1  # 从下一行开始
    # 向下读取直到遇到空行或另一个学历标题（简单处理：最多10行）
    for offset in range(10):
        current_row = row_start + offset
        field_cell = worksheet.cell(current_row, 1)
        if field_cell.value is None:
            break
        field_name = str(field_cell.value).strip()
        # 在 edu_data 中查找匹配的键
        for data_key, data_val in edu_data.items():
            if data_val and data_key == field_name:
                target_cell = worksheet.cell(current_row, 2)
                target_cell.value = data_val
                break

def fill_work_experience(worksheet, work_list):
    # 1. 定位关键字行
    keyword_row = find_row_with_keyword(worksheet, "工作经历（由近及远，仅限IT相关经历）")
    if not keyword_row:
        return
    # 2. 表头行 = keyword_row + 1
    header_row = keyword_row + 1
    # 3. 数据起始行 = header_row + 1
    data_start_row = header_row + 1
    # 4. 假设预留3行，我们先读取当前数据行区域（data_start_row 到 data_start_row+2）
    reserved_rows = 3
    current_data_end = data_start_row + reserved_rows - 1
    # 获取最大列数（表头列数）
    max_col = worksheet.max_column
    # 清空原有数据
    clear_rows(worksheet, data_start_row, current_data_end, max_col)
    
    # 5. 按开始日期倒序排序（由近及远）
    sorted_work = sorted(work_list, key=lambda x: x.get("开始日期", "1900-01-01"), reverse=True)
    
    # 6. 动态写入
    need_rows = len(sorted_work)
    if need_rows > reserved_rows:
        # 需要插入额外行
        insert_rows(worksheet, current_data_end + 1, need_rows - reserved_rows, max_col)
    elif need_rows < reserved_rows:
        # 多余行已经清空，无需额外操作
        pass
    
    for idx, work in enumerate(sorted_work):
        target_row = data_start_row + idx
        # 按表头顺序定位列：表头行的值决定列号
        # 表头行（header_row）各列内容：工作开始日期、工作结束日期、单位名称、岗位
        for col in range(1, max_col + 1):
            header_val = worksheet.cell(header_row, col).value
            if not header_val:
                continue
            if "开始日期" in str(header_val):
                worksheet.cell(target_row, col).value = normalize_date(work.get("开始日期", ""))
            elif "结束日期" in str(header_val):
                worksheet.cell(target_row, col).value = normalize_date(work.get("结束日期", ""))
            elif "单位名称" in str(header_val):
                worksheet.cell(target_row, col).value = work.get("单位名称", "")
            elif "岗位" in str(header_val):
                worksheet.cell(target_row, col).value = work.get("岗位", "")

def fill_project_experience(worksheet, project_list):
    # 类似工作经历，关键字为“项目经历（与上述工作经历匹配，仅IT相关经历）”
    keyword_row = find_row_with_keyword(worksheet, "项目经历（与上述工作经历匹配，仅IT相关经历）")
    if not keyword_row:
        return
    header_row = keyword_row + 1
    data_start_row = header_row + 1
    reserved_rows = 6
    current_data_end = data_start_row + reserved_rows - 1
    max_col = worksheet.max_column
    clear_rows(worksheet, data_start_row, current_data_end, max_col)
    
    sorted_proj = sorted(project_list, key=lambda x: x.get("开始日期", "1900-01-01"), reverse=True)
    need_rows = len(sorted_proj)
    if need_rows > reserved_rows:
        insert_rows(worksheet, current_data_end + 1, need_rows - reserved_rows, max_col)
    
    for idx, proj in enumerate(sorted_proj):
        target_row = data_start_row + idx
        for col in range(1, max_col + 1):
            header_val = worksheet.cell(header_row, col).value
            if not header_val:
                continue
            if "开始日期" in str(header_val):
                worksheet.cell(target_row, col).value = normalize_date(proj.get("开始日期", ""))
            elif "结束日期" in str(header_val):
                worksheet.cell(target_row, col).value = normalize_date(proj.get("结束日期", ""))
            elif "项目名称" in str(header_val):
                worksheet.cell(target_row, col).value = proj.get("项目名称", "")
            elif "项目描述" in str(header_val):
                worksheet.cell(target_row, col).value = proj.get("项目描述", "")
            elif "项目角色" in str(header_val):
                worksheet.cell(target_row, col).value = proj.get("项目角色", "")

def fill_template(template_path, output_path, ai_data):
    """主填充函数"""
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # 1. 基础信息
    fill_basic_info(ws, ai_data.get("basic", {}))
    
    # 2. 本科学历
    under = ai_data.get("education", {}).get("undergraduate", {})
    if under:
        fill_education_block(ws, "本科学历", under)
    
    # 3. 研究生学历
    post = ai_data.get("education", {}).get("postgraduate", {})
    if post:
        fill_education_block(ws, "研究生学历", post)
    
    # 4. 工作经历
    work_list = ai_data.get("work_experience", [])
    if work_list:
        fill_work_experience(ws, work_list)
    
    # 5. 项目经历
    proj_list = ai_data.get("project_experience", [])
    if proj_list:
        fill_project_experience(ws, proj_list)
    
    wb.save(output_path)
